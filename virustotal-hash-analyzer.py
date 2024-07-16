import openpyxl
import requests
import json
import logging

# Constants
API_KEY = 'd4123820878fe2baafb78f86d049083dee698a818a0ef2a6206dc98028917458'
EXCEL_FILE = 'file_hashes.xlsx'
SHEET_NAME = 'Sheet1'
VT_API_URL = 'https://www.virustotal.com/vtapi/v2/file/report'
OUTPUT_JSON_FILE = 'vt_results.json'
LOG_FILE = 'vt_errors.log'
OUTPUT_EXCEL_FILE = 'vt_results.xlsx'

# Configure logging
logging.basicConfig(filename=LOG_FILE, level=logging.ERROR,
                    format='%(asctime)s:%(levelname)s:%(message)s')

def get_file_report(file_hash):
    params = {'apikey': API_KEY, 'resource': file_hash}
    response = requests.get(VT_API_URL, params=params)
    if response.status_code == 200:
        try:
            return response.json()
        except ValueError:
            error_msg = f"Invalid JSON response: {response.text}"
            logging.error(f"Error fetching report for hash {file_hash}: {error_msg}")
            return {'error': error_msg}
    else:
        error_msg = f"HTTP {response.status_code} - {response.text}"
        logging.error(f"Error fetching report for hash {file_hash}: {error_msg}")
        return {'error': error_msg}

def read_hashes_from_excel(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    hashes = []
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        hashes.append(row[0])
    return hashes

def write_results_to_excel(results, output_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'VirusTotal Results'

    # Define the header
    headers = ['File Hash', 'Scan Date', 'Positives', 'Total', 'Permalink']
    sheet.append(headers)

    # Fill in the data
    for result in results:
        for file_hash, report in result.items():
            if 'error' in report:
                sheet.append([file_hash, 'Error', report['error'], '', ''])
            elif isinstance(report, dict):
                scan_date = report.get('scan_date', 'N/A')
                positives = report.get('positives', 'N/A')
                total = report.get('total', 'N/A')
                permalink = report.get('permalink', 'N/A')
                sheet.append([file_hash, scan_date, positives, total, permalink])
            else:
                # Handle unexpected response format
                logging.error(f"Unexpected report format for hash {file_hash}: {report}")
                sheet.append([file_hash, 'Error', 'Unexpected response format', '', ''])

    # Save the Excel file
    workbook.save(output_file)

def main():
    file_hashes = read_hashes_from_excel(EXCEL_FILE, SHEET_NAME)
    results = []
    for file_hash in file_hashes:
        report = get_file_report(file_hash)
        if 'error' in report:
            logging.error(f"Error fetching report for hash: {file_hash} - {report['error']}")
        results.append({file_hash: report})
    
    # Save results to JSON file
    with open(OUTPUT_JSON_FILE, 'w') as json_file:
        json.dump(results, json_file, indent=4)

    # Save results to Excel file
    write_results_to_excel(results, OUTPUT_EXCEL_FILE)

    print(f"Results saved to {OUTPUT_JSON_FILE} and {OUTPUT_EXCEL_FILE}")

if __name__ == "__main__":
    main()
